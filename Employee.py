import openpyxl,sqlite3



def addRecords():
	
	Emp_id=checkRecords()  # To check if records are already present in Database.
	
	wb=openpyxl.load_workbook('Employee.xlsx') # create workbook object
	sheet=wb.active.title			   # get the name of Active sheet in workbook
	conn=sqlite3.connect('Employee.db')	   # Connect to sqlite3 Employee database	
	
	if Emp_id!=0 and Emp_id!=wb[sheet].cell(row=wb[sheet].max_row,column=1).value:	# Check if records are present and last record in database table is not the last record of Employee.xlsx
		
		for row in range(2,wb[sheet].max_row+1):
			
			if Emp_id==wb[sheet].cell(row=row,column=1).value and row!=wb[sheet].max_row:
				new_row=row+1	# to traverse from row next to last record present in Employee Table
				# To store records in Employee Table
				for row in range(new_row,wb[sheet].max_row+1):
					
					sql='INSERT INTO Employee values ({0},"{1}",{2})'.format(wb[sheet].cell(row=row,column=1).value,wb[sheet].cell(row=row,column=2).value,wb[sheet].cell(row=row,column=3).value)
				
					conn.execute(sql)
					conn.commit()
	
	
	elif Emp_id==wb[sheet].cell(row=wb[sheet].max_row,column=1).value:	# Check if last record stored in Employee Table is same as last record present in Employee.xlsx
		print('Records already present!!')
								
	else:			# if no record is present in database table then store all data from Excel file into Employee table
		
		for row in range(2,wb[sheet].max_row+1):
			
			sql='INSERT INTO Employee values ({0},"{1}",{2})'.format(wb[sheet].cell(row=row,column=1).value,wb[sheet].cell(row=row,column=2).value,wb[sheet].cell(row=row,column=3).value)
				
			conn.execute(sql)
			conn.commit()
		
# Method checks the last record stored in Employee table and returns its Emp_id,If Employee table has no records then 0 is returned		
def checkRecords():
	Emp_id=0
	conn=sqlite3.connect('Employee.db')
	
	sql='select Emp_id from Employee ORDER BY Emp_id DESC LIMIT 1'
	result=conn.execute(sql)
	for emp_id in result:
		Emp_id=emp_id[0]
	
	return Emp_id


addRecords()
